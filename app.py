import os
import sys
import io
import json
import uuid
import threading
import webbrowser
import requests
from flask import Flask, render_template, request, jsonify, session, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.utils import secure_filename
from concurrent.futures import ThreadPoolExecutor, as_completed
import tempfile


def resource_path(relative_path):
    """获取资源的绝对路径，兼容 PyInstaller 打包后的路径"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


app = Flask(__name__, template_folder=resource_path('templates'))
app.secret_key = os.urandom(24)

UPLOAD_FOLDER = os.path.join(tempfile.gettempdir(), 'url_size_checker_uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'txt'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def read_urls_from_xlsx(file_path):
    urls = []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row and row[0]:
            url = str(row[0]).strip()
            if url.startswith(('http://', 'https://')):
                urls.append(url)
    wb.close()
    return urls


def read_urls_from_xls(file_path):
    import xlrd
    urls = []
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    for row_idx in range(ws.nrows):
        cell_value = ws.cell_value(row_idx, 0)
        if cell_value:
            url = str(cell_value).strip()
            if url.startswith(('http://', 'https://')):
                urls.append(url)
    wb.release_resources()
    return urls


def read_urls_from_txt(file_path):
    urls = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            url = line.strip()
            if url and url.startswith(('http://', 'https://')):
                urls.append(url)
    return urls


def create_session():
    """创建带连接池的 Session，复用 TCP/TLS 连接"""
    s = requests.Session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=20,
        pool_maxsize=20,
        max_retries=0,
    )
    s.mount('http://', adapter)
    s.mount('https://', adapter)
    return s


def get_file_size(url, session, retry=1):
    for attempt in range(retry + 1):
        try:
            resp = session.head(url, timeout=(3, 5), allow_redirects=True)
            cl = resp.headers.get('Content-Length')
            if cl:
                return int(cl), 'success'

            resp = session.get(url, timeout=(3, 5), stream=True, allow_redirects=True)
            cl = resp.headers.get('Content-Length')
            resp.close()
            if cl:
                return int(cl), 'success'

            return None, 'no_size'
        except requests.exceptions.Timeout:
            if attempt < retry:
                continue
            return None, 'timeout'
        except requests.exceptions.ConnectionError:
            if attempt < retry:
                continue
            return None, 'conn_error'
        except requests.exceptions.RequestException:
            if attempt < retry:
                continue
            return None, 'req_error'
        except Exception:
            if attempt < retry:
                continue
            return None, 'error'
    return None, 'failed'


def format_size(size):
    if size is None or size == 0:
        return '0 B'
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size < 1024:
            if unit == 'B':
                return f"{size} B"
            return f"{size:.2f} {unit}"
        size /= 1024
    return f"{size:.2f} PB"


STATUS_MAP = {
    'success': '✓ 成功',
    'no_size': '✗ 无大小信息',
    'timeout': '✗ 超时',
    'conn_error': '✗ 连接失败',
    'req_error': '✗ 请求失败',
    'error': '✗ 错误',
    'failed': '✗ 失败',
}


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': '不支持的文件格式，请上传 .xlsx、.xls 或 .txt 文件'}), 400

    filename = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4().hex}_{filename}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    file.save(file_path)

    try:
        ext = os.path.splitext(filename)[1].lower()
        if ext == '.xlsx':
            urls = read_urls_from_xlsx(file_path)
        elif ext == '.xls':
            urls = read_urls_from_xls(file_path)
        elif ext == '.txt':
            urls = read_urls_from_txt(file_path)
        else:
            return jsonify({'error': '不支持的文件格式'}), 400
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400
    finally:
        os.remove(file_path)

    if not urls:
        return jsonify({'error': '文件中未找到有效的URL（需以 http:// 或 https:// 开头）'}), 400

    return jsonify({'urls': urls, 'count': len(urls)})


@app.route('/check', methods=['POST'])
def check():
    data = request.get_json()
    if not data or 'urls' not in data:
        return jsonify({'error': '无URL数据'}), 400

    urls = data['urls']
    results = []
    session = create_session()

    with ThreadPoolExecutor(max_workers=20) as executor:
        future_to_idx = {
            executor.submit(get_file_size, url, session): (i, url)
            for i, url in enumerate(urls)
        }

        for future in as_completed(future_to_idx):
            idx, url = future_to_idx[future]
            size, status = future.result()
            results.append({
                'index': idx,
                'url': url,
                'size': size,
                'size_formatted': format_size(size) if size else '-',
                'status': status,
                'status_text': STATUS_MAP.get(status, '✗ 未知'),
            })

    results.sort(key=lambda r: (r['size'] is None, -(r['size'] or 0)))

    total_size = sum(r['size'] for r in results if r['size'] is not None)
    success_count = sum(1 for r in results if r['size'] is not None)
    fail_count = len(results) - success_count

    return jsonify({
        'results': results,
        'summary': {
            'total': len(results),
            'success': success_count,
            'fail': fail_count,
            'total_size': total_size,
            'total_size_formatted': format_size(total_size),
        }
    })


@app.route('/download', methods=['POST'])
def download():
    data = request.get_json()
    if not data or 'results' not in data:
        return jsonify({'error': '无数据'}), 400

    results = data['results']
    summary = data.get('summary', {})

    wb = Workbook()
    ws = wb.active
    ws.title = '文件大小检测结果'

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # 合计行
    ws.merge_cells('A1:D1')
    total_cell = ws['A1']
    total_cell.value = f"文件总大小合计: {summary.get('total_size_formatted', '-')}    (共 {summary.get('total', 0)} 个URL，成功 {summary.get('success', 0)} 个，失败 {summary.get('fail', 0)} 个)"
    total_cell.font = Font(bold=True, size=14, color='2563EB')
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # 表头
    headers = ['序号', 'URL', '文件大小', '状态']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # 数据行
    even_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    for i, r in enumerate(results):
        row = i + 3
        ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=r.get('url', ''))
        ws.cell(row=row, column=3, value=r.get('size_formatted', '-')).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=4, value=r.get('status_text', '')).alignment = Alignment(horizontal='center')
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = even_fill

    # 列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='文件大小检测结果.xlsx'
    )


def find_running_instance(start=5001, end=5099):
    """检测是否已有实例在运行，如果有则返回其端口号，否则返回 None"""
    import socket
    import urllib.request
    for port in range(start, end):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(0.5)
            result = sock.connect_ex(('127.0.0.1', port))
            sock.close()
            if result == 0:
                # 端口有服务在监听，验证是否是本应用
                try:
                    req = urllib.request.urlopen(
                        f'http://127.0.0.1:{port}/', timeout=2
                    )
                    # 检查返回内容是否包含本应用的特征
                    content = req.read(4096).decode('utf-8', errors='ignore')
                    if 'URL' in content or '文件大小' in content:
                        return port
                except Exception:
                    continue
        except Exception:
            continue
    return None


def find_free_port(start=5001, end=5099):
    """找一个可用的端口，找不到返回 None"""
    import socket
    for port in range(start, end):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.bind(('127.0.0.1', port))
                return port
        except OSError:
            continue
    return None


def wait_and_open_browser(port, max_wait=15):
    """等待服务就绪后再打开浏览器"""
    import time
    import socket
    url = f'http://127.0.0.1:{port}'
    for _ in range(max_wait * 10):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(0.5)
            result = sock.connect_ex(('127.0.0.1', port))
            sock.close()
            if result == 0:
                time.sleep(0.3)
                webbrowser.open(url)
                return
        except Exception:
            pass
        time.sleep(0.1)
    # 超时也尝试打开
    try:
        webbrowser.open(url)
    except Exception:
        pass


if __name__ == '__main__':
    import logging
    import signal
    import traceback

    # 写日志到临时目录，方便排查
    log_file = os.path.join(tempfile.gettempdir(), 'url_size_checker.log')
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s %(levelname)s %(message)s'
    )

    # 同时输出到 stderr（windowed 模式下也能写到日志）
    console_handler = logging.StreamHandler(sys.stderr)
    console_handler.setLevel(logging.DEBUG)
    logging.getLogger().addHandler(console_handler)

    def cleanup_and_exit(signum=None, frame=None):
        """确保进程干净退出"""
        logging.info("收到退出信号，正在关闭...")
        os._exit(0)

    signal.signal(signal.SIGINT, cleanup_and_exit)
    signal.signal(signal.SIGTERM, cleanup_and_exit)

    try:
        # 检测是否已有实例在运行
        existing_port = find_running_instance()
        if existing_port is not None:
            logging.info(f"检测到已有实例运行在端口 {existing_port}，直接打开浏览器")
            print(f"应用已在运行（端口 {existing_port}），正在打开浏览器...")
            webbrowser.open(f'http://127.0.0.1:{existing_port}')
            sys.exit(0)

        port = find_free_port()
        if port is None:
            error_msg = "无法找到可用端口（5001-5099 均被占用）"
            logging.error(error_msg)
            print(f"错误: {error_msg}")
            sys.exit(1)

        logging.info(f"使用端口: {port}")
        print(f"应用启动中，使用端口 {port}...")
        print(f"日志文件: {log_file}")
        print("关闭此窗口即可退出应用")

        # 后台线程：等服务真正启动后再打开浏览器
        threading.Thread(target=wait_and_open_browser, args=(port,), daemon=True).start()

        app.run(debug=False, port=port, host='127.0.0.1')
    except SystemExit:
        pass
    except Exception as e:
        error_msg = traceback.format_exc()
        logging.error(f"启动失败: {error_msg}")
        # windowed 模式下弹出错误对话框，让用户看到报错
        try:
            import subprocess
            subprocess.run([
                'osascript', '-e',
                f'display dialog "URL文件大小计算器启动失败:\\n{str(e)}\\n\\n日志文件: {log_file}" buttons {{"确定"}} default button "确定" with title "启动错误"'
            ], timeout=30)
        except Exception:
            pass
        sys.exit(1)
