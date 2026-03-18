import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import requests
import threading
from openpyxl import load_workbook
import os
import platform


class FileSizeCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("URL文件大小计算器")
        self.root.geometry("900x650")
        self.root.resizable(True, True)
        self.root.minsize(700, 500)

        self.urls = []
        self.results = []
        self.is_processing = False

        # 根据操作系统选择字体
        if platform.system() == 'Darwin':
            self.font_family = 'Helvetica Neue'
            self.mono_font = 'Menlo'
        elif platform.system() == 'Windows':
            self.font_family = 'Segoe UI'
            self.mono_font = 'Consolas'
        else:
            self.font_family = 'DejaVu Sans'
            self.mono_font = 'DejaVu Sans Mono'

        self.setup_ui()
        self.apply_styles()

    def apply_styles(self):
        style = ttk.Style()
        style.theme_use('clam')

        style.configure('Primary.TButton',
                        background='#2563EB',
                        foreground='white',
                        font=(self.font_family, 10),
                        padding=(20, 8))

        style.configure('Treeview',
                        font=(self.font_family, 10),
                        rowheight=32,
                        background='#FFFFFF',
                        fieldbackground='#FFFFFF',
                        foreground='#1E293B')

        style.configure('Treeview.Heading',
                        font=(self.font_family, 11, 'bold'),
                        background='#E2E8F0',
                        foreground='#1E293B')

        style.map('Treeview',
                  background=[('selected', '#DBEAFE')],
                  foreground=[('selected', '#1E293B')])

        style.configure('Progress.TProgressbar',
                        thickness=8,
                        troughcolor='#E2E8F0',
                        background='#2563EB')

        # 斑马纹通过 tag 实现
        self.tree.tag_configure('oddrow', background='#FFFFFF')
        self.tree.tag_configure('evenrow', background='#F8FAFC')
        self.tree.tag_configure('success', foreground='#16A34A')
        self.tree.tag_configure('fail', foreground='#DC2626')

    def setup_ui(self):
        self.root.configure(bg='#F8FAFC')

        # 顶部工具栏
        toolbar = tk.Frame(self.root, bg='#F8FAFC', height=60)
        toolbar.pack(fill=tk.X, padx=10, pady=10)

        self.btn_select = tk.Button(
            toolbar,
            text="📁 选择文件",
            font=(self.font_family, 10),
            bg='#2563EB',
            fg='white',
            activebackground='#1D4ED8',
            activeforeground='white',
            padx=16,
            pady=8,
            borderwidth=0,
            cursor='hand2',
            command=self.select_file
        )
        self.btn_select.pack(side=tk.LEFT)

        self.lbl_path = tk.Label(
            toolbar,
            text="请选择文件 (Excel / TXT)...",
            font=(self.font_family, 9),
            bg='#F8FAFC',
            fg='#64748B',
            anchor='w'
        )
        self.lbl_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=15)

        # 表格区域
        table_frame = tk.Frame(self.root, bg='#FFFFFF', padx=2, pady=2,
                               highlightbackground='#E2E8F0', highlightthickness=1)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        columns = ('index', 'url', 'size', 'status')
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings', selectmode='browse')

        self.tree.heading('index', text='序号')
        self.tree.heading('url', text='文件URL')
        self.tree.heading('size', text='文件大小')
        self.tree.heading('status', text='状态')

        self.tree.column('index', width=60, anchor='center', minwidth=50)
        self.tree.column('url', width=500, anchor='w', minwidth=200)
        self.tree.column('size', width=120, anchor='center', minwidth=80)
        self.tree.column('status', width=100, anchor='center', minwidth=70)

        # 垂直滚动条
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        # 水平滚动条
        h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # 底部状态栏
        status_frame = tk.Frame(self.root, bg='#F8FAFC', height=80)
        status_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        self.lbl_total = tk.Label(
            status_frame,
            text="文件总数: 0",
            font=(self.font_family, 10),
            bg='#F8FAFC',
            fg='#1E293B'
        )
        self.lbl_total.pack(side=tk.LEFT, padx=10)

        self.lbl_success = tk.Label(
            status_frame,
            text="成功: 0",
            font=(self.font_family, 10),
            bg='#F8FAFC',
            fg='#16A34A'
        )
        self.lbl_success.pack(side=tk.LEFT, padx=10)

        self.lbl_fail = tk.Label(
            status_frame,
            text="失败: 0",
            font=(self.font_family, 10),
            bg='#F8FAFC',
            fg='#DC2626'
        )
        self.lbl_fail.pack(side=tk.LEFT, padx=10)

        self.lbl_size = tk.Label(
            status_frame,
            text="总大小: 0 B",
            font=(self.font_family, 11, 'bold'),
            bg='#F8FAFC',
            fg='#2563EB'
        )
        self.lbl_size.pack(side=tk.LEFT, padx=20)

        self.progress = ttk.Progressbar(
            status_frame,
            mode='determinate',
            length=200,
            style='Progress.TProgressbar'
        )
        self.progress.pack(side=tk.RIGHT, padx=10)

        self.lbl_progress = tk.Label(
            status_frame,
            text="",
            font=(self.font_family, 9),
            bg='#F8FAFC',
            fg='#64748B'
        )
        self.lbl_progress.pack(side=tk.RIGHT, padx=5)

    def select_file(self):
        if self.is_processing:
            messagebox.showwarning("警告", "正在处理中，请稍候...")
            return

        file_path = filedialog.askopenfilename(
            title="选择文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("文本文件", "*.txt"), ("所有文件", "*.*")]
        )

        if file_path:
            self.lbl_path.config(text=file_path, fg='#1E293B')
            self.process_file(file_path)

    def process_file(self, file_path):
        self.clear_results()

        try:
            ext = os.path.splitext(file_path)[1].lower()

            if ext == '.xlsx':
                self._read_xlsx(file_path)
            elif ext == '.xls':
                self._read_xls(file_path)
            elif ext == '.txt':
                self._read_txt(file_path)
            else:
                messagebox.showerror("错误", "不支持的文件格式，请选择 .xlsx、.xls 或 .txt 文件")
                return

            if not self.urls:
                messagebox.showinfo("提示", "未在文件中找到有效的URL")
                return

            self.lbl_total.config(text=f"文件总数: {len(self.urls)}")
            self.start_processing()

        except ImportError as e:
            if 'xlrd' in str(e):
                messagebox.showerror("错误", "读取 .xls 文件需要安装 xlrd 库\n请运行: pip install xlrd")
            else:
                messagebox.showerror("错误", f"缺少依赖库: {str(e)}")
        except Exception as e:
            messagebox.showerror("错误", f"读取文件失败: {str(e)}")

    def _read_xlsx(self, file_path):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            if row and row[0]:
                url = str(row[0]).strip()
                if url and url.startswith(('http://', 'https://')):
                    self.urls.append(url)

        wb.close()

    def _read_xls(self, file_path):
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        for row_idx in range(ws.nrows):
            cell_value = ws.cell_value(row_idx, 0)
            if cell_value:
                url = str(cell_value).strip()
                if url and url.startswith(('http://', 'https://')):
                    self.urls.append(url)

        wb.release_resources()

    def _read_txt(self, file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                url = line.strip()
                if url and url.startswith(('http://', 'https://')):
                    self.urls.append(url)

    def start_processing(self):
        self.is_processing = True
        self.btn_select.config(state='disabled', bg='#94A3B8')
        self.progress['value'] = 0
        self.progress.pack(side=tk.RIGHT, padx=10)
        self.lbl_progress.config(text="处理中...")

        thread = threading.Thread(target=self.process_urls, daemon=True)
        thread.start()

    def process_urls(self):
        total = len(self.urls)
        success_count = 0
        fail_count = 0
        total_size = 0

        for i, url in enumerate(self.urls):
            size, status = self.get_file_size(url)

            self.results.append({
                'url': url,
                'size': size,
                'status': status
            })

            if size is not None:
                success_count += 1
                total_size += size
            else:
                fail_count += 1

            self.root.after(0, self.update_ui, i + 1, url, size, status,
                            success_count, fail_count, total_size, total)

        self.root.after(0, self.processing_complete)

    def get_file_size(self, url, retry=1):
        for attempt in range(retry + 1):
            try:
                # 先尝试 HEAD 请求
                response = requests.head(url, timeout=10, allow_redirects=True)
                content_length = response.headers.get('Content-Length')

                if content_length:
                    return int(content_length), '✓ 成功'

                # HEAD 无 Content-Length，尝试 GET（stream 模式，不下载内容）
                response = requests.get(url, timeout=10, stream=True, allow_redirects=True)
                content_length = response.headers.get('Content-Length')
                response.close()

                if content_length:
                    return int(content_length), '✓ 成功'

                return None, '✗ 无大小信息'

            except requests.exceptions.Timeout:
                if attempt < retry:
                    continue
                return None, '✗ 超时'
            except requests.exceptions.ConnectionError:
                if attempt < retry:
                    continue
                return None, '✗ 连接失败'
            except requests.exceptions.RequestException:
                if attempt < retry:
                    continue
                return None, '✗ 请求失败'
            except Exception:
                if attempt < retry:
                    continue
                return None, '✗ 错误'

        return None, '✗ 失败'

    def update_ui(self, index, url, size, status, success_count, fail_count, total_size, total):
        # 斑马纹
        row_tag = 'evenrow' if index % 2 == 0 else 'oddrow'

        # 显示完整 URL（水平滚动条可以查看）
        size_text = self.format_size(size) if size else '-'

        self.tree.insert('', tk.END, values=(
            index,
            url,
            size_text,
            status
        ), tags=(row_tag,))

        # 自动滚动到最新行
        children = self.tree.get_children()
        if children:
            self.tree.see(children[-1])

        self.lbl_success.config(text=f"成功: {success_count}")
        self.lbl_fail.config(text=f"失败: {fail_count}")
        self.lbl_size.config(text=f"总大小: {self.format_size(total_size)}")
        self.progress['value'] = (index / total) * 100
        self.lbl_progress.config(text=f"{index}/{total}")

    def processing_complete(self):
        self.is_processing = False
        self.btn_select.config(state='normal', bg='#2563EB')
        self.lbl_progress.config(text="完成")

        success = sum(1 for r in self.results if r['size'] is not None)
        total_size = sum(r['size'] for r in self.results if r['size'] is not None)

        messagebox.showinfo(
            "完成",
            f"处理完成！\n"
            f"共 {len(self.urls)} 个URL\n"
            f"成功获取: {success} 个\n"
            f"失败: {len(self.urls) - success} 个\n"
            f"文件总大小: {self.format_size(total_size)}"
        )

    def clear_results(self):
        self.urls = []
        self.results = []

        for item in self.tree.get_children():
            self.tree.delete(item)

        self.lbl_total.config(text="文件总数: 0")
        self.lbl_success.config(text="成功: 0")
        self.lbl_fail.config(text="失败: 0")
        self.lbl_size.config(text="总大小: 0 B")
        self.progress['value'] = 0
        self.lbl_progress.config(text="")

    @staticmethod
    def format_size(size):
        if size is None or size == 0:
            return '0 B'

        for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
            if size < 1024:
                if unit == 'B':
                    return f"{size} B"
                return f"{size:.2f} {unit}"
            size /= 1024
        return f"{size:.2f} PB"


def main():
    root = tk.Tk()
    FileSizeCheckerApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
